from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.models import User
from django.urls import reverse_lazy
from django.views.generic import ListView
from django.views.generic.edit import CreateView, DeleteView
from .models import InventoryList, SalesRecord
from .forms import SalesRecordForm
from django.db.models import Sum, F
from django.utils.timezone import now
from datetime import timedelta
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from django.utils.timezone import now
from datetime import timedelta
from .models import SalesRecord




# All links need review
# Create your views here.
class InventoryListView(LoginRequiredMixin, ListView):
    model = InventoryList
    template_name = 'main_g/inv_list.html'
    context_object_name = 'inventory'

    def get_queryset(self):
        return InventoryList.objects.filter(user=self.request.user)



class CreateInventoryView(CreateView):
    model = InventoryList
    fields = [
        'name', 'price', 'quantity'
    ]
    template_name = "main_g/create_inv.html"
    success_url = reverse_lazy("inventory")

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class DeleteInventoryView(DeleteView):
    model = InventoryList
    template_name = "main_g/delete_inv.html"
    success_url = reverse_lazy("inventory")

    def get_queryset(self):
        return  InventoryList.objects.filter(user=self.request.user)


class CreateSalesView(CreateView, ListView):
    model = SalesRecord
    form_class = SalesRecordForm
    template_name = "main_g/create_sales.html"
    success_url = reverse_lazy("sale_create")
    context_object_name = "sales"

    def get_queryset(self):
       return SalesRecord.objects.filter(user=self.request.user).order_by('-sale_date')[:5]

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user

        return kwargs

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class SalesListView(LoginRequiredMixin, ListView):
    model = SalesRecord
    template_name = "main_g/sales_list.html"
    context_object_name = "sales"

    def get_queryset(self):
       return SalesRecord.objects.filter(user=self.request.user).order_by('-sale_date')[:25]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        today = now().date()
        start_of_week = today - timedelta(days=today.weekday())  # Monday
        start_of_month = today.replace(day=1)

        qs = SalesRecord.objects.filter(user=self.request.user)

        context['totals'] = {
            'today': qs.filter(sale_date__date=today)
                     .aggregate(total=Sum(F('quantity_sold') * F('item__price')))['total'] or 0,

            'week': qs.filter(sale_date__date__gte=start_of_week)
                    .aggregate(total=Sum(F('quantity_sold') * F('item__price')))['total'] or 0,

            'month': qs.filter(sale_date__date__gte=start_of_month)
                     .aggregate(total=Sum(F('quantity_sold') * F('item__price')))['total'] or 0,
        }

        return context


def export_sales(request, period):
    today = now().date()
    wb = Workbook()

    def add_total_row(ws, start_row):
        """Adds bold total row at the bottom of the sheet"""
        total_row = ws.max_row + 1
        ws[f"B{total_row}"] = "TOTAL"
        ws[f"B{total_row}"].font = Font(bold=True)
        # If 'Total' is in column E

        ws[f"E{total_row}"] = f"=SUM(E{start_row}:E{total_row-1})"
        ws[f"E{total_row}"].font = Font(bold=True)

    if period == 'day':
        ws = wb.active
        ws.title = today.strftime("%Y-%m-%d")
        ws.append(["ID No.", "Item", "Quantity Sold", "Price", "Total", "Date"])
        start_row = 2
        sales = SalesRecord.objects.filter(user=request.user, sale_date__date=today)
        for sale in sales:
            ws.append([
                sale.sale_id,
                sale.item.name,
                sale.quantity_sold,
                float(sale.item.price),
                float(sale.total_sale_amount()),
                sale.sale_date.strftime("%Y-%m-%d %H:%M")
            ])
        add_total_row(ws, start_row)


    elif period == 'week':
        ws = wb.active
        ws.title = f"Week of {today.strftime('%Y-%m-%d')}"
        ws.append(["Date", "ID No." , "Item", "Quantity Sold", "Price", "Total"])

        start_row = 2
        start_of_week = today - timedelta(days=today.weekday())

        for i in range(7):
            day = start_of_week + timedelta(days=i)
            day_sales = SalesRecord.objects.filter(user=request.user, sale_date__date=day)
            if not day_sales.exists():
                continue

            # Day header
            ws.append([day.strftime("%A, %Y-%m-%d")])
            day_start_row = ws.max_row + 1

            for sale in day_sales:
                ws.append([

                    "",  # Empty date cell for alignment
                    sale.sale_id,
                    sale.item.name,
                    sale.quantity_sold,
                    float(sale.item.price),
                    float(sale.total_sale_amount())

                ])

            # Subtotal for the day

            subtotal_row = ws.max_row + 1

            ws[f"C{subtotal_row}"] = "Subtotal"

            ws[f"C{subtotal_row}"].font = Font(bold=True)

            ws[f"F{subtotal_row}"] = f"=SUM(F{day_start_row}:{'F' + str(subtotal_row - 1)})"

            ws[f"F{subtotal_row}"].font = Font(bold=True)

        # Grand total

        final_row = ws.max_row + 1

        ws[f"C{final_row}"] = "TOTAL"

        ws[f"C{final_row}"].font = Font(bold=True)

        ws[f"F{final_row}"] = f"=SUM(F2:F{final_row - 1})"

        ws[f"F{final_row}"].font = Font(bold=True)


    elif period == 'month':

        start_of_month = today.replace(day=1)

        first_monday = start_of_month - timedelta(days=start_of_month.weekday())

        current_week_start = first_monday

        while current_week_start.month <= today.month:

            ws = wb.create_sheet(title=f"Week {current_week_start.strftime('%W')}")

            ws.append(["Date", "ID No." , "Item", "Quantity Sold", "Price", "Total"])

            start_row = 2

            for i in range(7):

                day = current_week_start + timedelta(days=i)

                if day.month != today.month:
                    continue

                day_sales = SalesRecord.objects.filter(user=request.user, sale_date__date=day)

                if not day_sales.exists():
                    continue

                # Day header

                ws.append([day.strftime("%A, %Y-%m-%d")])

                day_start_row = ws.max_row + 1

                for sale in day_sales:
                    ws.append([

                        "",  # Empty date cell for alignment
                        sale.sale_id,
                        sale.item.name,
                        sale.quantity_sold,
                        float(sale.item.price),
                        float(sale.total_sale_amount())

                    ])

                # Subtotal for the day

                subtotal_row = ws.max_row + 1

                ws[f"C{subtotal_row}"] = "Subtotal"

                ws[f"C{subtotal_row}"].font = Font(bold=True)

                ws[f"F{subtotal_row}"] = f"=SUM(F{day_start_row}:{'F' + str(subtotal_row - 1)})"

                ws[f"F{subtotal_row}"].font = Font(bold=True)

            # Grand total for the week

            final_row = ws.max_row + 1

            ws[f"C{final_row}"] = "TOTAL"

            ws[f"C{final_row}"].font = Font(bold=True)

            ws[f"F{final_row}"] = f"=SUM(F2:F{final_row - 1})"

            ws[f"F{final_row}"].font = Font(bold=True)

            current_week_start += timedelta(days=7)

    # Prepare HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"sales_{period}_{today}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response







